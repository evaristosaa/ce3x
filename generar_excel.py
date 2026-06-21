#!/usr/bin/env python3
"""
Generador del Excel plantilla CE3X
Crea el fichero CE3X_Inmuebles.xlsx con todas las columnas necesarias
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paleta de colores ──────────────────────────────────────────────────────────
COLOR_HEADER      = "1F4E79"   # azul oscuro
COLOR_AUTO        = "D9E1F2"   # azul claro — relleno automático (Catastro)
COLOR_REQUIRED    = "FFF2CC"   # amarillo  — obligatorio manual
COLOR_OPTIONAL    = "F2F2F2"   # gris claro — opcional
COLOR_STATUS_OK   = "E2EFDA"   # verde
COLOR_STATUS_PEND = "FCE4D6"   # naranja

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_AUTO   = Font(name="Calibri", color="1F4E79", size=10)
FONT_NORMAL = Font(name="Calibri", size=10)

def header_style(ws, cell_ref, text):
    c = ws[cell_ref]
    c.value = text
    c.font  = FONT_HEADER
    c.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_style(ws, cell_ref):
    c = ws[cell_ref]
    c.fill = PatternFill("solid", fgColor=COLOR_AUTO)
    c.font = FONT_AUTO

def req_style(ws, cell_ref):
    c = ws[cell_ref]
    c.fill = PatternFill("solid", fgColor=COLOR_REQUIRED)
    c.font = FONT_NORMAL

def opt_style(ws, cell_ref):
    c = ws[cell_ref]
    c.fill = PatternFill("solid", fgColor=COLOR_OPTIONAL)
    c.font = FONT_NORMAL

# ── Columnas ───────────────────────────────────────────────────────────────────
COLUMNS = [
    # (header, ancho, tipo: A=auto, R=required, O=optional)
    # GESTIÓN
    ("ID",                          6,  "O"),
    ("Estado",                      20, "R"),
    ("Fecha Alta",                  14, "O"),
    ("Técnico",                     20, "O"),
    ("Observaciones",               30, "O"),

    # IDENTIFICACIÓN — auto desde Catastro
    ("Ref. Catastral",              24, "R"),
    ("Provincia",                   16, "A"),
    ("Municipio",                   18, "A"),
    ("Dirección",                   40, "A"),
    ("CP",                           8, "A"),
    ("Uso",                         16, "A"),
    ("Tipo Inmueble",               20, "R"),   # unifamiliar/bloque/local/otro
    ("Sup. Construida (m²)",        18, "A"),
    ("Año Construcción",            16, "A"),

    # DATOS GEOMÉTRICOS
    ("Sup. Útil (m²)",              16, "R"),
    ("Nº Plantas",                  12, "R"),
    ("Altura libre (m)",            16, "O"),
    ("Orientación fachada ppal",    22, "R"),

    # ENVOLVENTE TÉRMICA
    ("Fachada: composición",        30, "R"),
    ("Fachada: espesor (cm)",       20, "R"),
    ("Fachada: U (W/m²K)",          18, "O"),
    ("Cubierta: tipo",              20, "R"),
    ("Cubierta: composición",       28, "R"),
    ("Cubierta: U (W/m²K)",         18, "O"),
    ("Suelo: tipo",                 20, "R"),
    ("Suelo: composición",          28, "O"),
    ("Huecos: tipo vidrio",         22, "R"),
    ("Huecos: tipo marco",          20, "R"),
    ("Huecos: % superficie",        18, "R"),
    ("Huecos: U vidrio (W/m²K)",    20, "O"),
    ("Huecos: factor solar",        18, "O"),

    # INSTALACIONES — calefacción
    ("Calef: tipo",                 20, "R"),
    ("Calef: combustible",          20, "R"),
    ("Calef: rendimiento (%)",      20, "R"),
    ("Calef: año instalación",      18, "O"),

    # INSTALACIONES — refrigeración
    ("Refrig: tipo",                20, "R"),
    ("Refrig: SEER/EER",            16, "O"),

    # INSTALACIONES — ACS
    ("ACS: tipo",                   20, "R"),
    ("ACS: combustible",            20, "R"),
    ("ACS: rendimiento (%)",        20, "R"),

    # INSTALACIONES — ventilación
    ("Ventilación: tipo",           20, "R"),
    ("Ventilación: caudal (m³/h)",  22, "O"),

    # SOLO TERCIARIO — iluminación
    ("Ilum: potencia (W/m²)",       20, "O"),
    ("Ilum: tipo lámpara",          20, "O"),

    # RENOVABLES
    ("Renovable: tipo",             20, "O"),
    ("Renovable: potencia/sup",     20, "O"),

    # RESULTADO
    ("Calificación",                14, "O"),
    ("Emisiones CO₂ (kgCO₂/m²·a)", 24, "O"),
    ("Consumo EP (kWh/m²·a)",       24, "O"),
    ("Fichero .ce3x",               30, "O"),
    ("Fichero PDF",                 30, "O"),
    ("Fichero XML",                 30, "O"),
    ("Fecha Certificado",           18, "O"),
    ("Nº Registro",                 20, "O"),
]

def create_excel(path):
    wb = openpyxl.Workbook()

    # ── Hoja principal ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Inmuebles"
    ws.freeze_panes = "G2"
    ws.row_dimensions[1].height = 45

    for col_idx, (col_name, col_width, col_type) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        header_style(ws, f"{col_letter}1", col_name)
        ws.column_dimensions[col_letter].width = col_width

    # Fila de ejemplo
    example = {
        "ID": 1,
        "Estado": "pendiente_info",
        "Fecha Alta": "2026-06-01",
        "Técnico": "Evaristo Saa",
        "Ref. Catastral": "0128501TG4302N0004BU",
        "Provincia": "Sevilla",
        "Municipio": "Dos Hermanas",
        "Dirección": "PL SEN-1 ENTRENUCLEOS 40(D) Es:4 Pl:00 Pt:01",
        "CP": "41704",
        "Uso": "Residencial",
        "Tipo Inmueble": "bloque",
        "Sup. Construida (m²)": 240,
        "Año Construcción": 2021,
    }

    col_names = [c[0] for c in COLUMNS]
    for col_idx, (col_name, _, col_type) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        cell_ref = f"{col_letter}2"
        val = example.get(col_name, "")
        ws[cell_ref] = val
        if col_type == "A":
            auto_style(ws, cell_ref)
        elif col_type == "R":
            req_style(ws, cell_ref)
        else:
            opt_style(ws, cell_ref)

    # Validación desplegable — Estado
    estado_col = col_names.index("Estado") + 1
    dv_estado = DataValidation(
        type="list",
        formula1='"pendiente_info,completada_info,generando,generada_documentacion,error,cancelada"',
        allow_blank=True,
        showDropDown=False
    )
    dv_estado.sqref = f"{get_column_letter(estado_col)}2:{get_column_letter(estado_col)}1000"
    ws.add_data_validation(dv_estado)

    # Validación desplegable — Tipo Inmueble
    tipo_col = col_names.index("Tipo Inmueble") + 1
    dv_tipo = DataValidation(
        type="list",
        formula1='"unifamiliar,bloque,local,oficina,nave,otro"',
        allow_blank=True,
        showDropDown=False
    )
    dv_tipo.sqref = f"{get_column_letter(tipo_col)}2:{get_column_letter(tipo_col)}1000"
    ws.add_data_validation(dv_tipo)

    # ── Hoja Leyenda ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Leyenda")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 50

    leyenda = [
        ("Color", "Significado"),
        ("🔵 Azul claro", "Relleno automático desde API Catastro — no editar"),
        ("🟡 Amarillo", "Campo obligatorio — rellenar manualmente"),
        ("⬜ Gris claro", "Campo opcional"),
    ]
    estados = [
        ("Estado", "Descripción"),
        ("pendiente_info", "Ref catastral introducida, faltan datos"),
        ("completada_info", "Todos los datos listos — mAI puede generar .ce3x"),
        ("generando", "mAI está procesando"),
        ("generada_documentacion", "PDF y XML generados ✅"),
        ("error", "Algo fue mal — ver Observaciones"),
        ("cancelada", "Descartado"),
    ]

    ws2["A1"] = "LEYENDA DE COLORES"
    ws2["A1"].font = FONT_HEADER
    ws2["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for i, (k, v) in enumerate(leyenda[1:], start=2):
        ws2[f"A{i}"] = k
        ws2[f"B{i}"] = v

    ws2["A7"] = "ESTADOS"
    ws2["A7"].font = FONT_HEADER
    ws2["A7"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for i, (k, v) in enumerate(estados[1:], start=8):
        ws2[f"A{i}"] = k
        ws2[f"B{i}"] = v
        if k == "generada_documentacion":
            ws2[f"A{i}"].fill = PatternFill("solid", fgColor=COLOR_STATUS_OK)
        elif k in ("pendiente_info", "error"):
            ws2[f"A{i}"].fill = PatternFill("solid", fgColor=COLOR_STATUS_PEND)

    wb.save(path)
    print(f"Excel generado: {path}")

if __name__ == "__main__":
    import os
    out = os.path.expanduser(
        "/mnt/c/Users/evari/Documents/CE3X_Inmuebles.xlsx"
    )
    create_excel(out)
