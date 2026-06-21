#!/usr/bin/env python3
"""
Small helper for CE3X_Gestion.xlsx.

For now it adds or updates a cadastral reference in the INMUEBLES sheet using
public Catastro data. Manual CE3X fields stay untouched.
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(__file__).with_name("CE3X_Gestion.xlsx")
CATASTRO_URL = (
    "https://ovc.catastro.meh.es/OVCServWeb/OVCSWLocalizacionRC/"
    "OVCCallejero.asmx/Consulta_DNPRC"
)
NS = {"cat": "http://www.catastro.meh.es/"}


@dataclass(frozen=True)
class CatastroData:
    ref_catastral: str
    direccion: str | None
    municipio: str | None
    provincia: str | None
    codigo_postal: str | None
    uso: str | None
    superficie_catastro: float | None
    ano_construccion: int | None


def text_at(root: ET.Element, path: str) -> str | None:
    node = root.find(path, NS)
    if node is None or node.text is None:
        return None
    value = node.text.strip()
    return value or None


def parse_number(value: str | None) -> float | None:
    if not value:
        return None
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return None


def parse_int(value: str | None) -> int | None:
    number = parse_number(value)
    return int(number) if number is not None else None


def query_catastro(ref_catastral: str) -> CatastroData:
    ref = ref_catastral.strip().upper().replace(" ", "")
    if len(ref) != 20:
        raise ValueError("La referencia catastral debe tener 20 caracteres.")

    query = urllib.parse.urlencode({"Provincia": "", "Municipio": "", "RC": ref})
    with urllib.request.urlopen(f"{CATASTRO_URL}?{query}", timeout=20) as response:
        body = response.read()

    root = ET.fromstring(body)
    error = text_at(root, ".//cat:lerr/cat:err/cat:des")
    if error:
        raise RuntimeError(f"Catastro devuelve error: {error}")

    bi = root.find(".//cat:bico/cat:bi", NS)
    if bi is None:
        raise RuntimeError("Catastro no ha devuelto datos del inmueble.")

    return CatastroData(
        ref_catastral=ref,
        direccion=text_at(bi, "cat:ldt"),
        municipio=text_at(bi, "cat:dt/cat:nm"),
        provincia=text_at(bi, "cat:dt/cat:np"),
        codigo_postal=text_at(bi, ".//cat:dp"),
        uso=text_at(bi, "cat:debi/cat:luso"),
        superficie_catastro=parse_number(text_at(bi, "cat:debi/cat:sfc")),
        ano_construccion=parse_int(text_at(bi, "cat:debi/cat:ant")),
    )


def header_map(ws) -> dict[str, int]:
    # Row 2 contains the real field names; row 1 is only section grouping.
    return {
        normalize_header(ws.cell(2, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(2, col).value
    }


def normalize_header(value: object) -> str:
    return " ".join(str(value).replace("²", "2").split())


def find_or_create_row(ws, ref: str, ref_col: int) -> tuple[int, bool]:
    normalized = ref.strip().upper()
    first_empty = None
    for row in range(3, ws.max_row + 1):
        current = ws.cell(row, ref_col).value
        if current and str(current).strip().upper() == normalized:
            return row, False
        if not current and first_empty is None:
            first_empty = row

    row = first_empty or max(ws.max_row + 1, 3)
    return row, True


def update_workbook(path: Path, data: CatastroData, *, save: bool) -> tuple[int, bool]:
    wb = load_workbook(path)
    ws = wb["INMUEBLES"]
    columns = header_map(ws)

    required = [
        "Ref. Catastral",
        "Estado",
        "Direccion",
        "Municipio",
        "Provincia",
        "C.Postal",
        "Uso",
        "Ano Const.",
        "Sup. Catastro (m2)",
        "Observaciones",
    ]
    # The workbook has accented labels. Keep tolerant aliases for script usage.
    aliases = {
        "Ref. Catastral": "Ref. Catastral",
        "Direccion": "Dirección",
        "Ano Const.": "Año Const.",
        "Sup. Catastro (m2)": "Sup. Catastro (m2)",
    }
    resolved = {
        key: columns.get(normalize_header(key)) or columns.get(normalize_header(aliases.get(key, "")))
        for key in required
    }
    missing = [key for key, col in resolved.items() if col is None]
    if missing:
        raise RuntimeError(f"Faltan columnas esperadas en INMUEBLES: {', '.join(missing)}")

    row, created = find_or_create_row(ws, data.ref_catastral, resolved["Ref. Catastral"])
    values = {
        "Ref. Catastral": data.ref_catastral,
        "Estado": "PENDIENTE" if created else ws.cell(row, resolved["Estado"]).value or "PENDIENTE",
        "Direccion": data.direccion,
        "Municipio": data.municipio,
        "Provincia": data.provincia.title() if data.provincia else None,
        "C.Postal": data.codigo_postal,
        "Uso": data.uso,
        "Ano Const.": data.ano_construccion,
        "Sup. Catastro (m2)": data.superficie_catastro,
    }

    for key, value in values.items():
        if value is not None:
            ws.cell(row, resolved[key]).value = value

    obs_cell = ws.cell(row, resolved["Observaciones"])
    stamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    note = f"Catastro actualizado {stamp}"
    obs_cell.value = f"{obs_cell.value}; {note}" if obs_cell.value else note

    if save:
        wb.save(path)

    return row, created


def main() -> int:
    parser = argparse.ArgumentParser(description="Actualiza CE3X_Gestion.xlsx desde Catastro.")
    parser.add_argument("ref_catastral", help="Referencia catastral de 20 caracteres.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Ruta del Excel. Por defecto: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument("--dry-run", action="store_true", help="Consulta y valida sin guardar.")
    args = parser.parse_args()

    data = query_catastro(args.ref_catastral)
    row, created = update_workbook(args.workbook, data, save=not args.dry_run)

    if args.dry_run:
        action = "crearia" if created else "actualizaria"
    else:
        action = "creada" if created else "actualizada"
    print(f"Fila {row} {action}: {data.ref_catastral}")
    print(f"- Direccion: {data.direccion}")
    print(f"- Municipio/Provincia: {data.municipio} / {data.provincia}")
    print(f"- CP: {data.codigo_postal}")
    print(f"- Uso: {data.uso}")
    print(f"- Superficie catastro: {data.superficie_catastro}")
    print(f"- Ano construccion: {data.ano_construccion}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
